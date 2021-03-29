#!/usr/bin/env python3
#coding:utf-8
import sys
import os
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from utils.logger import Logger

ROW_KEY_IDX = 2
ROW_START_IDX = 4

logger = Logger(sys.stderr)

workbook_cache = {}
def load_workbook_data_only(file_path):
	workbook = None
	if file_path not in workbook_cache:
		workbook = load_workbook(file_path, read_only=True, data_only = True)
		workbook_cache[file_path] = workbook
	else:
		workbook = workbook_cache[file_path]
	return workbook

def close_cache_file(file_path):
	if not workbook_cache.has_key(file_path):
		return
	workbook_cache[file_path].close()
	workbook_cache.pop(file_path)

def convert_key_map_2_key_col(sheet, key_map):
	key_col_map = {}
	row_idx = 0
	for row in sheet.rows:
		row_idx = row_idx + 1
		if row_idx == ROW_KEY_IDX:
			for cel_idx in range(len(row)):
				cell = row[cel_idx]
				key = cell.value
				if key in key_map:
					key_col_map[key] = cel_idx+1
			break

	return key_col_map

def read_excel_data(file_path, sheet_name, key_map, main_key = None):
	start_t = time.time()
	excel_file = load_workbook_data_only(file_path)
	sheet = excel_file[sheet_name]
	key_col_map = convert_key_map_2_key_col(sheet, key_map)
	data_map = {}
	main_value = 0

	row_idx = 0
	for row in sheet.rows:
		row_idx = row_idx + 1
		if row_idx < ROW_START_IDX:
			continue
		data = {"row_idx" : row_idx}
		for key in key_map.keys():
			if key not in key_col_map:
				continue
			col_idx = key_col_map[key]
			value = row[col_idx - 1].value
			if value == None:
				value = ""
			data[key] = value

		if main_key:
			main_value = data[main_key]
		else:
			main_value = main_value + 1

		assert main_value not in data_map, "第%d行:key(%s)value(%s)重复了! 文件(%s)表(%s)"%(row_idx, str(main_key), str(main_value), file_path, sheet_name)

		data_map[main_value] = data

	return data_map

def read_excel_texts(file_path, sheet_name, key_map):
	start_t = time.time()
	excel_file = load_workbook_data_only(file_path)
	sheet = excel_file[sheet_name]
	key_col_map = convert_key_map_2_key_col(sheet, key_map)
	texts_map = {}
	for key in key_map:
		texts_map[key] = []

	row_idx = 0
	for row in sheet.rows:
		row_idx = row_idx + 1
		if row_idx < ROW_START_IDX:
			continue

		for key in key_map.keys():
			if key not in key_col_map:
				continue
			col_idx = key_col_map[key]
			value = row[col_idx - 1].value
			if value == None:
				continue
			if type(value) != str:
				value = str(value)
			texts_map[key].append(value)

	return  texts_map


def write_excel_data(file_path, sheet_name, row_data_map):
	start_t = time.time()
	# logger.info("write_excel_data start:%s,%s"%(file_path, sheet_name))
	excel_file = load_workbook(file_path)
	sheet = excel_file[sheet_name]
	ncols = sheet.max_column

	key_col_map = {}
	for col_idx in range(1, ncols + 1):
		key = sheet.cell(row = ROW_KEY_IDX, column = col_idx).value
		if key:
			key_col_map[key] = col_idx

	for row_idx, row_data in row_data_map.items():
		for key in row_data.keys():
			col_idx = key_col_map[key]
			if col_idx:
				value = row_data[key]
				sheet.cell(row = row_idx, column = col_idx, value = value)
	excel_file.save(filename = file_path)

