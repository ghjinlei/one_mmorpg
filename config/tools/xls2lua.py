#!/usr/bin/env python3
#coding:utf-8
import sys
import os
import re
import argparse
from sys import exit
from openpyxl import load_workbook
sys.path.append(os.path.dirname(__file__))
from utils.logger import Logger

KEY_ROW   = 2
TYPE_ROW  = 3
DATA_ROW  = 4

CONTENT = "content"

FIELD_OP_VALID    = "invalid"
FIELD_OP_NAME     = "name"
FIELD_OP_TYPE     = "type"
FIELD_OP_LIST     = "list"
FIELD_OP_SERVER   = "s"
FIELD_OP_CLIENT   = "c"

VTYPE_RAWDATA     = "rawdata"
VTYPE_INT         = "int"
VTYPE_FLOAT       = "float"
VTYPE_BOOL        = "bool"
VTYPE_BIT         = "bit"
VTYPE_STRING      = "string"

logger = Logger(sys.stderr)

pre_dump_table = None
def load_hookfile(filename):
    import imp
    global pre_dump_table
    try:
        module = imp.load_source('d', filename)
        if hasattr(module, 'pre_dump_table'):
            pre_dump_table = module.pre_dump_table
    except:
        pass

class ExtendValue:
    def __init__(self, value, comment=None):
        self.value = value
        self.comment = comment
    def __str__(self):
        vtype = type(self.value)
        if vtype == type(True):
            return vtype and "true" or "false"
        return str(self.value)

def xls_format(cell, sheet_name):
    value = cell.value
    # TODO: 特殊类型处理
    return value

def adjust_type(value):
    if isinstance(value, float) and round(value) == value:
        return int(value)
    return value

def parse_type(value, vtype):
    if type(value) == type(""):
        value = value.strip()
    if value == "" or value == None:
        return None

    if vtype == VTYPE_RAWDATA:
        value = adjust_type(value)
        value = ExtendValue(value)
    elif vtype == VTYPE_INT:
        try:
            value = int(value)
        except:
            exit("'%s' can not convert to int." % value)
    elif vtype == VTYPE_FLOAT:
        try:
            value = float(value)
        except:
            exit("'%s' can not convert to float." % value)
    elif vtype == VTYPE_BOOL:
        try:
            value = bool(value)
        except:
            exit("'%s' can not convert to bool." % value)
    elif vtype == VTYPE_STRING:
        value = str(adjust_type(value))
    elif vtype == VTYPE_BIT:
        bitstrlist = str(adjust_type(value)).split(",")
        value = 0
        for i in range(0, len(bitstrlist)):
            value = value + math.pow(2, int(bitstrlist[i]))
    elif re.match("func\((.*)\)", vtype):
        if type(value) == type(True):
            value = value and "true" or "false"
        m = re.match("func\((.*)\)", vtype)
        args = m.group(1)
        value = "function(%s) return %s end" % (args, value)
        value = ExtentType(value)

    return value

def parse_value(value, field_info):
    if field_info[FIELD_OP_LIST]:
        if value != None:
            value = value.split("|")
            for i in range(len(value)):
                value = parse_type(value, field_info[FIELD_OP_TYPE])
    else:
        value = parse_type(value, field_info[FIELD_OP_TYPE])

    return value

def try_convert_int(key):
    try:
        return int(key)
    except:
        return key

def push_value(value, path, data_table):
    path_list = path.split(".")
    cur_table = data_table
    for path in path_list[:-1]:
        if not path in cur_table:
            path = try_convert_int(path)
            cur_table[path] = {}
        cur_table = cur_table[path]

    key = try_convert_int(path_list[-1])
    cur_table[key] = value

def parse_sheet(main_table, ws, keyrow, typerow, datarow, sheet_name, export_type):
    maxcol = ws.max_column
    maxrow = ws.max_row
    logger.info("parse_sheet,sheet_name=%s,maxcol=%d, maxrow=%d"%(sheet_name, maxcol, maxrow))

    field_info_dict = {}
    # 处理字段信息
    for col_idx in range(1, maxcol + 1):
        keycell = ws.cell(row=keyrow, column=col_idx)
        vtypecell = ws.cell(row=typerow, column=col_idx)
        key = xls_format(keycell, sheet_name)
        vtype = xls_format(vtypecell, sheet_name)

        field_info = {
                FIELD_OP_VALID    : True,
                FIELD_OP_NAME     : None,
                FIELD_OP_TYPE     : None,
                FIELD_OP_LIST     : False,
                FIELD_OP_SERVER   : False,
                FIELD_OP_CLIENT   : False,
        }
        field_info_dict[col_idx] = field_info

        if key is None or key.startswith("//"):
            field_info[FIELD_OP_VALID] = False
            continue

        field_info[FIELD_OP_NAME] = key.strip()

        type_list = vtype.split("|")
        for i in range(len(type_list)):
            type_key = type_list[i]
            if type_key == FIELD_OP_LIST:
                field_info[FIELD_OP_LIST]
            if type_key == FIELD_OP_SERVER:
                field_info[FIELD_OP_SERVER] = True
            elif type_key ==  FIELD_OP_SERVER:
                field_info[FIELD_OP_SERVER] = True
            else:
                field_info[FIELD_OP_TYPE] = type_key

    sheet_table = {}
    for row_idx in range(datarow, maxrow + 1):
        row_data = {}
        recordidcell = ws.cell(row=row_idx, column=1)
        recordid = xls_format(recordidcell, sheet_name)
        if not recordid:
            break

        sheet_table[recordid] = row_data
        for col_idx in range(1, maxcol + 1):
            field_info = field_info_dict[col_idx]
            if not field_info[FIELD_OP_VALID]:
                continue
            if field_info[FIELD_OP_CLIENT] and export_type == FIELD_OP_SERVER:
                continue
            if field_info[FIELD_OP_SERVER] and export_type == FIELD_OP_CLIENT:
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            value = xls_format(cell, sheet_name)
            value = parse_value(value, field_info)
            if value is None:
                continue
            push_value(value, field_info[FIELD_OP_NAME], row_data)

    return sheet_table


def gen_table(xlspath, export_type):
    wb = load_workbook(filename=xlspath, data_only=True)
    ws_list = wb.worksheets

    main_ws = ws_list[0]
    assert main_ws.title == "main", "第一张工作表必须是main(%s)" % xlspath

    main_table = {}
    main_table = parse_sheet(main_table, main_ws, KEY_ROW, TYPE_ROW, DATA_ROW, "main", export_type)

    for ws in ws_list:
        sheet_name = ws.title

        if sheet_name not in main_table:
            continue

        output = None
        if "output" in main_table[sheet_name]:
            output = main_table[sheet_name]["output"]
        output = "sc" if not output else output

        if output.find(export_type) > -1:
            sheet_table = parse_sheet(main_table, ws, KEY_ROW, TYPE_ROW, DATA_ROW, sheet_name, export_type)
            main_table[sheet_name][CONTENT] = sheet_table

    return main_table

DictType = type({})
IntType = type(1)
FloatType = type(1.0)
BooleanType = type(True)
StringType = type("")
NoneType = type(None)
ListType = type([])
ExtendType = type(ExtendValue(""))

base_type_dict = {
    IntType       : True,
    FloatType     : True,
    BooleanType   : True,
    StringType    : True,
    NoneType      : True,
    ListType      : True,
    ExtendType    : True,
}
def dump_base_type(value, f):
    write = f.write
    value_type = type(value)
    if value_type == IntType:
        write("%d" % value)
    elif value_type == FloatType:
        float_str = "%f" % value
        m = re.match("(\-*\d+\.\d*?)(0*)$", float_str)
        value = m.group(1).rstrip(".")
        write(value)
    elif value_type == BooleanType:
        value = value == True and "true" or "false"
        write(value)
    elif value_type == StringType:
        write("[=[%s]=]" % value)
    elif value_type == NoneType:
        write("nil")
    elif value_type == ListType:
        write("{")
        for x in value:
            dump_base_type(x, f)
            write(",")
        write("}")
    elif value_type == InstanceType:
        write(str(value))

def dump_value(value, level=1, f=None):
    f = f if f else sys.stdout
    write = f.write

    value_type = type(value)
    if value_type in base_type_dict:
        dump_base_type(value, f)
    elif value_type == DictType:
        write("{\n")

        items = list(value.items())
        items.sort()
        for k, v in items:
            for i in range(level):
                write("\t")
            if type(k) == IntType:
                write("[%d] = " % k)
            else:
                write("%s = " % k)
            dump_value(v, level + 1, f)
            write(",\n")

        for i in range(level - 1):
            write("\t")
        write("}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename", help="the input filename or filepath")
    parser.add_argument("-t", dest="export_type", help="export type:s/c")
    parser.add_argument("-i", dest="hookfile", help="hook file")
    args = parser.parse_args()

    filename = args.filename
    export_type = args.export_type or FIELD_OP_SERVER

    if not os.path.isfile(filename):
        exit("no input file specified")

    data_table = gen_table(filename, export_type)

    if pre_dump_table:
        data_table = pre_dump_table(data_table, filename)

    output_f = sys.stdout
    output_f.write("local datatable =\n")
    dump_value(data_table, 1, output_f)
    output_f.write("\nfunction get_datatable() return datatable end\n")
    output_f.write("\nfunction get_content(sheet_name) return datatable[sheet_name].content end\n")

if __name__  == "__main__":
    main()
